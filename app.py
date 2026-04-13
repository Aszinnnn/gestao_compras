from flask import Flask, render_template, request, jsonify, send_file, make_response
from flask_cors import CORS
import pandas as pd
import numpy as np
import os
import base64
import io
import matplotlib
matplotlib.use('Agg')
import matplotlib.pyplot as plt
from werkzeug.utils import secure_filename
import re
import chardet
from datetime import datetime
import json
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.formatting.rule import ColorScaleRule, CellIsRule, FormulaRule
from openpyxl.chart import BarChart, Reference, PieChart

app = Flask(__name__)
CORS(app)
app.config['UPLOAD_FOLDER'] = 'uploads'
app.config['MAX_CONTENT_LENGTH'] = 50 * 1024 * 1024
app.config['SECRET_KEY'] = 'sua-chave-secreta-super-segura-2024'

os.makedirs(app.config['UPLOAD_FOLDER'], exist_ok=True)
os.makedirs('static/css', exist_ok=True)
os.makedirs('static/js', exist_ok=True)

ALLOWED_EXTENSIONS = {'csv'}

def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS

def detect_encoding(file_path):
    with open(file_path, 'rb') as f:
        raw_data = f.read()
        result = chardet.detect(raw_data)
        return result['encoding'] or 'utf-8'

class AnalisadorCompras:
    def __init__(self, df):
        self.df = df.copy()
        
    def limpar_valor_monetario(self, valor):
        if pd.isna(valor):
            return 0.0
        valor_str = str(valor).strip()
        if isinstance(valor, (int, float)) and not isinstance(valor, bool):
            return float(valor)
        valor_str = re.sub(r'R?\$?\s*', '', valor_str)
        valor_str = re.sub(r'[^\d,\-\.]', '', valor_str)
        if ',' in valor_str and '.' in valor_str:
            if valor_str.rfind(',') > valor_str.rfind('.'):
                valor_str = valor_str.replace('.', '')
                valor_str = valor_str.replace(',', '.')
        elif ',' in valor_str:
            valor_str = valor_str.replace(',', '.')
        valor_str = valor_str.replace('.', '')
        try:
            return float(valor_str) if valor_str else 0.0
        except:
            return 0.0
    
    def limpar_dados(self):
        colunas_padrao = {
            'codigo': ['sku', 'codigo', 'código', 'id_produto', 'product_id', 'id', 'cod', 'item_code', 'produto_codigo'],
            'descricao': ['descricao', 'produto', 'nome', 'description', 'product_name', 'desc', 'item', 'product'],
            'quantidade': ['quantidade', 'qtd', 'quantity', 'qty', 'quant', 'amount'],
            'preco': ['preco', 'preço', 'valor', 'price', 'unit_price', 'valor_unitario', 'preco_unitario'],
            'fornecedor': ['fornecedor', 'vendedor', 'supplier', 'vendor', 'forn', 'provider', 'seller'],
            'prazo': ['prazo', 'lead_time', 'dias_entrega', 'delivery_days', 'entrega', 'leadtime'],
            'frete': ['frete', 'shipping', 'freight', 'custo_frete', 'freight_cost'],
            'categoria': ['categoria', 'category', 'departamento', 'tipo', 'type']
        }
        
        for padrao, variacoes in colunas_padrao.items():
            for col in self.df.columns:
                if col.lower() in variacoes:
                    self.df.rename(columns={col: padrao}, inplace=True)
                    break
        
        if 'preco' not in self.df.columns:
            raise ValueError(f"Coluna de preço não encontrada. Colunas: {list(self.df.columns)}")
        
        self.df['preco'] = self.df['preco'].apply(self.limpar_valor_monetario)
        
        if 'frete' in self.df.columns:
            self.df['frete'] = self.df['frete'].apply(self.limpar_valor_monetario)
        else:
            self.df['frete'] = 0
        
        if 'quantidade' in self.df.columns:
            self.df['quantidade'] = pd.to_numeric(self.df['quantidade'], errors='coerce').fillna(1)
        else:
            self.df['quantidade'] = 1
        
        if 'prazo' in self.df.columns:
            self.df['prazo'] = pd.to_numeric(self.df['prazo'], errors='coerce').fillna(15).astype(int)
        else:
            self.df['prazo'] = 15
        
        if 'descricao' not in self.df.columns:
            self.df['descricao'] = self.df['codigo'] if 'codigo' in self.df.columns else 'Produto'
        
        if 'fornecedor' not in self.df.columns:
            self.df['fornecedor'] = 'Fornecedor Padrão'
        
        if 'codigo' not in self.df.columns:
            self.df['codigo'] = [f'PROD_{i}' for i in range(len(self.df))]
        
        if 'categoria' not in self.df.columns:
            self.df['categoria'] = 'Geral'
        
        # Calcular custo total corretamente
        self.df['custo_total'] = (self.df['preco'] * self.df['quantidade']) + self.df['frete']
        self.df = self.df.dropna(subset=['preco'])
        self.df = self.df[self.df['preco'] > 0]
        self.df = self.df[self.df['quantidade'] > 0]
        
        return self.df
    
    def funil_melhores_precos(self):
        try:
            funil = self.df.loc[self.df.groupby('codigo')['preco'].idxmin()]
            funil = funil[['codigo', 'descricao', 'categoria', 'fornecedor', 'preco', 'quantidade', 'prazo', 'frete', 'custo_total']]
            funil = funil.sort_values('preco')
            return funil.head(50).to_dict('records')
        except:
            return []
    
    def comparacao_fornecedores(self):
        try:
            comparacao = []
            for codigo in self.df['codigo'].unique()[:20]:
                codigo_data = self.df[self.df['codigo'] == codigo]
                if len(codigo_data) > 1:
                    fornecedores = []
                    for _, row in codigo_data.iterrows():
                        fornecedores.append({
                            'fornecedor': str(row['fornecedor']),
                            'preco': round(float(row['preco']), 2),
                            'prazo': int(row['prazo']),
                            'custo_total': round(float(row['custo_total']), 2)
                        })
                    comparacao.append({
                        'codigo': str(codigo),
                        'descricao': str(codigo_data.iloc[0]['descricao']),
                        'fornecedores': fornecedores,
                        'economia_possivel': round(float(codigo_data['preco'].max() - codigo_data['preco'].min()), 2)
                    })
            return comparacao
        except:
            return []
    
    def analise_fornecedores(self):
        try:
            analise = self.df.groupby('fornecedor').agg({
                'codigo': 'count',
                'preco': ['mean', 'min', 'max'],
                'custo_total': 'sum',
                'prazo': 'mean'
            }).round(2)
            analise.columns = ['total_itens', 'preco_medio', 'preco_minimo', 'preco_maximo', 'valor_total', 'prazo_medio']
            analise = analise.sort_values('valor_total', ascending=False)
            resultado = analise.reset_index().head(20).to_dict('records')
            for item in resultado:
                for key in item:
                    if isinstance(item[key], (np.integer, np.int64)):
                        item[key] = int(item[key])
                    elif isinstance(item[key], (np.floating, np.float64)):
                        item[key] = float(item[key])
                item['fornecedor'] = str(item['fornecedor'])
                item['prazo_medio'] = int(round(item['prazo_medio']))  # Prazo como inteiro
            return resultado
        except:
            return []
    
    def analise_por_categoria(self):
        try:
            analise = self.df.groupby('categoria').agg({
                'codigo': 'count',
                'preco': 'mean',
                'custo_total': 'sum'
            }).round(2)
            analise.columns = ['total_itens', 'preco_medio', 'valor_total']
            analise = analise.sort_values('valor_total', ascending=False)
            resultado = analise.reset_index().to_dict('records')
            for item in resultado:
                item['valor_total'] = float(item['valor_total'])
                item['preco_medio'] = float(item['preco_medio'])
            return resultado
        except:
            return []
    
    def compras_fragmentadas(self):
        try:
            fragmentadas = self.df.groupby('codigo')['fornecedor'].nunique()
            fragmentadas = fragmentadas[fragmentadas > 1]
            resultado = []
            for codigo, num_forn in fragmentadas.head(20).items():
                dados = self.df[self.df['codigo'] == codigo].iloc[0]
                resultado.append({
                    'codigo': str(codigo),
                    'descricao': str(dados['descricao']),
                    'num_fornecedores': int(num_forn),
                    'sugestao': f"Consolidar compras em um único fornecedor pode gerar desconto"
                })
            return resultado
        except:
            return []
    
    def outliers_preco(self):
        try:
            stats = self.df.groupby('codigo')['preco'].agg(['mean', 'std'])
            stats = stats[stats['std'] > 0]
            stats['limite_superior'] = stats['mean'] + (2 * stats['std'])
            outliers = []
            for codigo in stats.index[:20]:
                codigo_data = self.df[self.df['codigo'] == codigo]
                limite = stats.loc[codigo, 'limite_superior']
                acima = codigo_data[codigo_data['preco'] > limite]
                for _, row in acima.iterrows():
                    outliers.append({
                        'codigo': str(codigo),
                        'descricao': str(row['descricao']),
                        'fornecedor': str(row['fornecedor']),
                        'preco': round(float(row['preco']), 2),
                        'media': round(float(stats.loc[codigo, 'mean']), 2),
                        'diferenca_percentual': round(float(((row['preco'] - stats.loc[codigo, 'mean']) / stats.loc[codigo, 'mean']) * 100), 1)
                    })
            return outliers
        except:
            return []
    
    def recomendacoes_cotacao(self):
        try:
            recomendacoes = []
            for codigo in self.df['codigo'].unique()[:20]:
                codigo_data = self.df[self.df['codigo'] == codigo]
                if len(codigo_data) >= 2:
                    preco_min = codigo_data['preco'].min()
                    preco_medio = codigo_data['preco'].mean()
                    if preco_medio > preco_min * 1.15:
                        recomendacoes.append({
                            'codigo': str(codigo),
                            'descricao': str(codigo_data.iloc[0]['descricao']),
                            'preco_atual_medio': round(float(preco_medio), 2),
                            'melhor_preco': round(float(preco_min), 2),
                            'economia_potencial': round(float(preco_medio - preco_min), 2),
                            'acao': f"Cotar para atingir R$ {preco_min:.2f}"
                        })
            return recomendacoes
        except:
            return []
    
    def estatisticas_gerais(self):
        try:
            economia_total = 0
            try:
                economia_total = float(self.df.groupby('codigo')['preco'].apply(lambda x: x.max() - x.min()).sum())
            except:
                economia_total = 0
            
            return {
                'total_itens': int(len(self.df)),
                'total_produtos': int(self.df['codigo'].nunique()),
                'total_fornecedores': int(self.df['fornecedor'].nunique()),
                'total_categorias': int(self.df['categoria'].nunique()),
                'valor_total_gasto': round(float(self.df['custo_total'].sum()), 2),
                'preco_medio_geral': round(float(self.df['preco'].mean()), 2),
                'economia_potencial_total': round(economia_total, 2),
                'produtos_com_multiplos_fornecedores': int(self.df.groupby('codigo')['fornecedor'].nunique()[lambda x: x > 1].count())
            }
        except:
            return {
                'total_itens': 0,
                'total_produtos': 0,
                'total_fornecedores': 0,
                'total_categorias': 0,
                'valor_total_gasto': 0,
                'preco_medio_geral': 0,
                'economia_potencial_total': 0,
                'produtos_com_multiplos_fornecedores': 0
            }
    
    def gerar_graficos_base64(self):
        graficos = {}
        
        try:
            fig, ax = plt.subplots(figsize=(10, 6))
            fig.patch.set_facecolor('#1a1a1a')
            ax.set_facecolor('#2d2d2d')
            
            top_forn = self.df.groupby('fornecedor')['custo_total'].sum().sort_values(ascending=False).head(10)
            if len(top_forn) > 0:
                cores = plt.cm.YlOrRd(range(len(top_forn)))
                ax.barh(range(len(top_forn)), top_forn.values, color=cores)
                ax.set_yticks(range(len(top_forn)))
                ax.set_yticklabels(top_forn.index, fontsize=10, color='white')
                ax.set_xlabel('Valor Total (R$)', fontsize=12, color='#FFD700')
                ax.set_title('Top 10 Fornecedores por Valor Gasto', fontsize=14, fontweight='bold', color='#FFD700')
                ax.invert_yaxis()
                ax.tick_params(colors='white')
                plt.tight_layout()
                
                buffer = io.BytesIO()
                plt.savefig(buffer, format='png', dpi=100, bbox_inches='tight', facecolor='#1a1a1a')
                buffer.seek(0)
                graficos['top_fornecedores'] = base64.b64encode(buffer.getvalue()).decode()
            plt.close(fig)
        except:
            graficos['top_fornecedores'] = ""
        
        # Outros gráficos...
        try:
            fig, ax = plt.subplots(figsize=(10, 6))
            fig.patch.set_facecolor('#1a1a1a')
            ax.set_facecolor('#2d2d2d')
            ax.hist(self.df['preco'], bins=30, color='#FFD700', edgecolor='white', alpha=0.7)
            ax.set_xlabel('Preço Unitário (R$)', fontsize=12, color='#FFD700')
            ax.set_ylabel('Frequência', fontsize=12, color='#FFD700')
            ax.set_title('Distribuição de Preços', fontsize=14, fontweight='bold', color='#FFD700')
            ax.grid(True, alpha=0.3, color='white')
            ax.tick_params(colors='white')
            plt.tight_layout()
            buffer = io.BytesIO()
            plt.savefig(buffer, format='png', dpi=100, bbox_inches='tight', facecolor='#1a1a1a')
            buffer.seek(0)
            graficos['distribuicao_precos'] = base64.b64encode(buffer.getvalue()).decode()
            plt.close(fig)
        except:
            graficos['distribuicao_precos'] = ""
        
        try:
            fig, ax = plt.subplots(figsize=(12, 6))
            fig.patch.set_facecolor('#1a1a1a')
            ax.set_facecolor('#2d2d2d')
            top_produtos = self.df.groupby('codigo')['preco'].mean().sort_values(ascending=False).head(15)
            if len(top_produtos) > 0:
                cores = plt.cm.YlOrRd(range(len(top_produtos)))
                ax.bar(range(len(top_produtos)), top_produtos.values, color=cores)
                ax.set_xticks(range(len(top_produtos)))
                ax.set_xticklabels(top_produtos.index, rotation=45, ha='right', fontsize=9, color='white')
                ax.set_xlabel('Código Produto', fontsize=12, color='#FFD700')
                ax.set_ylabel('Preço Médio (R$)', fontsize=12, color='#FFD700')
                ax.set_title('Top 15 Produtos por Preço Médio', fontsize=14, fontweight='bold', color='#FFD700')
                ax.grid(True, alpha=0.3, axis='y', color='white')
                ax.tick_params(colors='white')
                plt.tight_layout()
                buffer = io.BytesIO()
                plt.savefig(buffer, format='png', dpi=100, bbox_inches='tight', facecolor='#1a1a1a')
                buffer.seek(0)
                graficos['top_produtos'] = base64.b64encode(buffer.getvalue()).decode()
            plt.close(fig)
        except:
            graficos['top_produtos'] = ""
        
        try:
            fig, ax = plt.subplots(figsize=(10, 6))
            fig.patch.set_facecolor('#1a1a1a')
            ax.set_facecolor('#2d2d2d')
            cat_data = self.df.groupby('categoria')['custo_total'].sum().sort_values(ascending=False).head(8)
            if len(cat_data) > 0:
                cores = plt.cm.YlOrRd(range(len(cat_data)))
                wedges, texts, autotexts = ax.pie(cat_data.values, labels=cat_data.index, autopct='%1.1f%%', colors=cores)
                for text in texts:
                    text.set_color('white')
                for autotext in autotexts:
                    autotext.set_color('white')
                ax.set_title('Distribuição por Categoria', fontsize=14, fontweight='bold', color='#FFD700')
                plt.tight_layout()
                buffer = io.BytesIO()
                plt.savefig(buffer, format='png', dpi=100, bbox_inches='tight', facecolor='#1a1a1a')
                buffer.seek(0)
                graficos['categorias'] = base64.b64encode(buffer.getvalue()).decode()
            plt.close(fig)
        except:
            graficos['categorias'] = ""
        
        return graficos

# Import numpy for type conversion
import numpy as np

@app.route('/')
def index():
    return render_template('index.html')

@app.route('/static/<path:path>')
def serve_static(path):
    return send_from_directory('static', path)

@app.route('/upload', methods=['POST'])
def upload_file():
    try:
        if 'file' not in request.files:
            return jsonify({'error': 'Nenhum arquivo enviado'}), 400
        
        file = request.files['file']
        if file.filename == '':
            return jsonify({'error': 'Nenhum arquivo selecionado'}), 400
        
        if not allowed_file(file.filename):
            return jsonify({'error': 'Tipo de arquivo não permitido. Use apenas CSV'}), 400
        
        filename = secure_filename(file.filename)
        filepath = os.path.join(app.config['UPLOAD_FOLDER'], filename)
        file.save(filepath)
        
        try:
            encoding = detect_encoding(filepath)
            separadores = [',', ';', '\t', '|']
            df = None
            
            for sep in separadores:
                try:
                    df = pd.read_csv(filepath, encoding=encoding, sep=sep)
                    if len(df.columns) > 1:
                        break
                except:
                    continue
            
            if df is None or len(df.columns) < 2:
                raise ValueError("Não foi possível ler o arquivo CSV")
            
            analisador = AnalisadorCompras(df)
            analisador.limpar_dados()
            
            resultado = {
                'estatisticas': analisador.estatisticas_gerais(),
                'funil': analisador.funil_melhores_precos(),
                'comparacao_fornecedores': analisador.comparacao_fornecedores(),
                'analise_fornecedores': analisador.analise_fornecedores(),
                'analise_categorias': analisador.analise_por_categoria(),
                'compras_fragmentadas': analisador.compras_fragmentadas(),
                'outliers': analisador.outliers_preco(),
                'recomendacoes': analisador.recomendacoes_cotacao(),
                'graficos': analisador.gerar_graficos_base64()
            }
            
            return jsonify(resultado)
            
        except Exception as e:
            import traceback
            traceback.print_exc()
            return jsonify({'error': f'Erro ao processar: {str(e)}'}), 500
        finally:
            if os.path.exists(filepath):
                try:
                    os.remove(filepath)
                except:
                    pass
    
    except Exception as e:
        return jsonify({'error': f'Erro no servidor: {str(e)}'}), 500

@app.route('/exportar/excel', methods=['POST'])
def exportar_excel():
    try:
        data = request.json
        if not data:
            return jsonify({'error': 'Sem dados para exportar'}), 400
        
        # Criar um buffer em memória
        output = io.BytesIO()
        
        # Criar Excel com múltiplas abas
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            # Aba 1: Funil de Preços
            df_funil = pd.DataFrame(data.get('funil', []))
            if not df_funil.empty:
                df_funil.to_excel(writer, sheet_name='Funil de Preços', index=False)
                
                # Formatar aba Funil
                worksheet = writer.sheets['Funil de Preços']
                
                # Cabeçalho bonito
                header_fill = PatternFill(start_color='FFD700', end_color='FFD700', fill_type='solid')
                header_font = Font(bold=True, color='000000', size=11)
                
                for col in range(1, len(df_funil.columns) + 1):
                    cell = worksheet.cell(row=1, column=col)
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                
                # Ajustar largura das colunas
                for column in worksheet.columns:
                    max_length = 0
                    column_letter = column[0].column_letter
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = min(max_length + 2, 30)
                    worksheet.column_dimensions[column_letter].width = adjusted_width
                
                # Formatar números
                for row in range(2, len(df_funil) + 2):
                    for col, col_name in enumerate(df_funil.columns, 1):
                        cell = worksheet.cell(row=row, column=col)
                        if 'preco' in col_name.lower() or 'custo' in col_name.lower() or 'total' in col_name.lower():
                            cell.number_format = 'R$ #,##0.00'
                        elif 'prazo' in col_name.lower():
                            cell.number_format = '0'
                
                # Adicionar filtros
                worksheet.auto_filter.ref = worksheet.dimensions
            
            # Aba 2: Análise Fornecedores
            df_fornecedores = pd.DataFrame(data.get('analise_fornecedores', []))
            if not df_fornecedores.empty:
                df_fornecedores.to_excel(writer, sheet_name='Análise Fornecedores', index=False)
                
                worksheet2 = writer.sheets['Análise Fornecedores']
                
                # Cabeçalho
                for col in range(1, len(df_fornecedores.columns) + 1):
                    cell = worksheet2.cell(row=1, column=col)
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                
                # Ajustar colunas
                for column in worksheet2.columns:
                    max_length = 0
                    column_letter = column[0].column_letter
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = min(max_length + 2, 25)
                    worksheet2.column_dimensions[column_letter].width = adjusted_width
                
                # Formatar números
                for row in range(2, len(df_fornecedores) + 2):
                    for col, col_name in enumerate(df_fornecedores.columns, 1):
                        cell = worksheet2.cell(row=row, column=col)
                        if 'preco' in col_name.lower() or 'valor' in col_name.lower():
                            cell.number_format = 'R$ #,##0.00'
                        elif 'prazo' in col_name.lower():
                            cell.number_format = '0'
                
                worksheet2.auto_filter.ref = worksheet2.dimensions
                
                # Adicionar formatação condicional (valores altos em vermelho)
                red_fill = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')
                valor_col = None
                for idx, col_name in enumerate(df_fornecedores.columns, 1):
                    if 'valor_total' in col_name.lower():
                        valor_col = idx
                        break
                
                if valor_col:
                    for row in range(2, len(df_fornecedores) + 2):
                        cell = worksheet2.cell(row=row, column=valor_col)
                        if cell.value and isinstance(cell.value, (int, float)) and cell.value > 50000:
                            cell.fill = red_fill
                            cell.font = Font(color='FFFFFF', bold=True)
            
            # Aba 3: Recomendações
            df_recomendacoes = pd.DataFrame(data.get('recomendacoes', []))
            if not df_recomendacoes.empty:
                df_recomendacoes.to_excel(writer, sheet_name='Recomendações', index=False)
                
                worksheet3 = writer.sheets['Recomendações']
                
                # Cabeçalho
                for col in range(1, len(df_recomendacoes.columns) + 1):
                    cell = worksheet3.cell(row=1, column=col)
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                
                # Ajustar colunas
                for column in worksheet3.columns:
                    max_length = 0
                    column_letter = column[0].column_letter
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = min(max_length + 2, 40)
                    worksheet3.column_dimensions[column_letter].width = adjusted_width
                
                # Formatar números
                for row in range(2, len(df_recomendacoes) + 2):
                    for col, col_name in enumerate(df_recomendacoes.columns, 1):
                        cell = worksheet3.cell(row=row, column=col)
                        if 'preco' in col_name.lower() or 'economia' in col_name.lower():
                            cell.number_format = 'R$ #,##0.00'
                
                worksheet3.auto_filter.ref = worksheet3.dimensions
            
            # Aba 4: Comparação Fornecedores (detalhada)
            df_comparacao = pd.DataFrame(data.get('comparacao_fornecedores', []))
            if not df_comparacao.empty:
                # Expandir a comparação para linhas individuais
                linhas_expandidas = []
                for item in df_comparacao.to_dict('records'):
                    for forn in item.get('fornecedores', []):
                        linhas_expandidas.append({
                            'Código': item['codigo'],
                            'Produto': item['descricao'],
                            'Fornecedor': forn['fornecedor'],
                            'Preço': forn['preco'],
                            'Prazo (dias)': forn['prazo'],
                            'Custo Total': forn['custo_total'],
                            'Economia Potencial': item['economia_possivel']
                        })
                
                df_comparacao_expandida = pd.DataFrame(linhas_expandidas)
                if not df_comparacao_expandida.empty:
                    df_comparacao_expandida.to_excel(writer, sheet_name='Comparação Completa', index=False)
                    
                    worksheet4 = writer.sheets['Comparação Completa']
                    
                    # Cabeçalho
                    for col in range(1, len(df_comparacao_expandida.columns) + 1):
                        cell = worksheet4.cell(row=1, column=col)
                        cell.fill = header_fill
                        cell.font = header_font
                        cell.alignment = Alignment(horizontal='center', vertical='center')
                    
                    # Ajustar colunas
                    for column in worksheet4.columns:
                        max_length = 0
                        column_letter = column[0].column_letter
                        for cell in column:
                            try:
                                if len(str(cell.value)) > max_length:
                                    max_length = len(str(cell.value))
                            except:
                                pass
                        adjusted_width = min(max_length + 2, 30)
                        worksheet4.column_dimensions[column_letter].width = adjusted_width
                    
                    # Formatar números
                    for row in range(2, len(df_comparacao_expandida) + 2):
                        for col, col_name in enumerate(df_comparacao_expandida.columns, 1):
                            cell = worksheet4.cell(row=row, column=col)
                            if 'Preço' in col_name or 'Custo' in col_name or 'Economia' in col_name:
                                cell.number_format = 'R$ #,##0.00'
                            elif 'Prazo' in col_name:
                                cell.number_format = '0'
                    
                    worksheet4.auto_filter.ref = worksheet4.dimensions
                    
                    # Formatação condicional para destacar melhores preços
                    for row in range(2, len(df_comparacao_expandida) + 2):
                        codigo_atual = worksheet4.cell(row=row, column=1).value
                        preco_atual = worksheet4.cell(row=row, column=4).value
                        
                        # Encontrar o menor preço para este código
                        precos_mesmo_codigo = []
                        for r in range(2, len(df_comparacao_expandida) + 2):
                            if worksheet4.cell(row=r, column=1).value == codigo_atual:
                                precos_mesmo_codigo.append(worksheet4.cell(row=r, column=4).value)
                        
                        if precos_mesmo_codigo and preco_atual == min(precos_mesmo_codigo):
                            green_fill = PatternFill(start_color='92D050', end_color='92D050', fill_type='solid')
                            for col in range(1, 8):
                                cell = worksheet4.cell(row=row, column=col)
                                cell.fill = green_fill
            
            # Aba 5: Resumo Executivo
            stats = data.get('estatisticas', {})
            df_resumo = pd.DataFrame([
                ['Métrica', 'Valor'],
                ['Total de Itens', stats.get('total_itens', 0)],
                ['Total de Produtos', stats.get('total_produtos', 0)],
                ['Total de Fornecedores', stats.get('total_fornecedores', 0)],
                ['Total de Categorias', stats.get('total_categorias', 0)],
                ['Valor Total Gasto', f"R$ {stats.get('valor_total_gasto', 0):,.2f}"],
                ['Preço Médio Geral', f"R$ {stats.get('preco_medio_geral', 0):,.2f}"],
                ['Economia Potencial Total', f"R$ {stats.get('economia_potencial_total', 0):,.2f}"],
                ['Produtos com Múltiplos Fornecedores', stats.get('produtos_com_multiplos_fornecedores', 0)]
            ])
            
            df_resumo.to_excel(writer, sheet_name='Resumo Executivo', index=False, header=False)
            
            worksheet5 = writer.sheets['Resumo Executivo']
            
            # Formatar resumo
            for row in range(1, 10):
                cell_metric = worksheet5.cell(row=row, column=1)
                cell_value = worksheet5.cell(row=row, column=2)
                
                if row == 1:
                    cell_metric.fill = header_fill
                    cell_value.fill = header_fill
                    cell_metric.font = header_font
                    cell_value.font = header_font
                else:
                    cell_metric.font = Font(bold=True, color='000000')
                    if row % 2 == 0:
                        cell_metric.fill = PatternFill(start_color='F0F0F0', end_color='F0F0F0', fill_type='solid')
                        cell_value.fill = PatternFill(start_color='F0F0F0', end_color='F0F0F0', fill_type='solid')
            
            worksheet5.column_dimensions['A'].width = 30
            worksheet5.column_dimensions['B'].width = 25
        
        output.seek(0)
        
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=f'Relatorio_Compras_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
        )
    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({'error': str(e)}), 500

@app.route('/exportar/csv', methods=['POST'])
def exportar_csv():
    try:
        data = request.json
        if not data:
            return jsonify({'error': 'Sem dados para exportar'}), 400
        
        # Criar CSV com todos os dados
        output = io.StringIO()
        
        # Funil de Preços
        output.write("=== FUNIL DE PREÇOS ===\n")
        df_funil = pd.DataFrame(data.get('funil', []))
        if not df_funil.empty:
            df_funil.to_csv(output, index=False, encoding='utf-8-sig')
        
        output.write("\n\n=== ANÁLISE FORNECEDORES ===\n")
        df_fornecedores = pd.DataFrame(data.get('analise_fornecedores', []))
        if not df_fornecedores.empty:
            df_fornecedores.to_csv(output, index=False, encoding='utf-8-sig')
        
        output.write("\n\n=== RECOMENDAÇÕES ===\n")
        df_recomendacoes = pd.DataFrame(data.get('recomendacoes', []))
        if not df_recomendacoes.empty:
            df_recomendacoes.to_csv(output, index=False, encoding='utf-8-sig')
        
        output.seek(0)
        
        response = make_response(output.getvalue())
        response.headers['Content-Type'] = 'text/csv; charset=utf-8'
        response.headers['Content-Disposition'] = f'attachment; filename=Relatorio_Compras_{datetime.now().strftime("%Y%m%d_%H%M%S")}.csv'
        
        return response
    except Exception as e:
        return jsonify({'error': str(e)}), 500

if __name__ == '__main__':
    app.run(debug=True, host='0.0.0.0', port=5000)